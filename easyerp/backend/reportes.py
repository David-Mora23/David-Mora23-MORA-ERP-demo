"""
reportes.py - Dashboard, KPIs, gráficos y exportación PDF/Excel.
"""

import io
from datetime import datetime
from flask import Blueprint, request, send_file
from database import get_db
from utils import token_required, rows_to_list, success_response

reportes_bp = Blueprint('reportes', __name__, url_prefix='/api')


@reportes_bp.route('/dashboard/kpis', methods=['GET'])
@token_required
def dashboard_kpis():
    """Retorna los KPIs principales del dashboard."""
    conn = get_db()

    total_ventas = conn.execute(
        "SELECT COALESCE(SUM(total), 0) as total FROM facturas WHERE estado != 'anulada'"
    ).fetchone()['total']

    num_facturas = conn.execute(
        "SELECT COUNT(*) as count FROM facturas WHERE estado != 'anulada'"
    ).fetchone()['count']

    total_productos = conn.execute('SELECT COUNT(*) as count FROM productos').fetchone()['count']

    stock_bajo = conn.execute(
        'SELECT COUNT(*) as count FROM productos WHERE stock <= 10'
    ).fetchone()['count']

    total_clientes = conn.execute('SELECT COUNT(*) as count FROM clientes').fetchone()['count']
    total_proveedores = conn.execute('SELECT COUNT(*) as count FROM proveedores').fetchone()['count']
    total_empleados = conn.execute(
        "SELECT COUNT(*) as count FROM empleados WHERE estado = 'activo'"
    ).fetchone()['count']

    ingresos = conn.execute(
        "SELECT COALESCE(SUM(monto), 0) as total FROM transacciones WHERE tipo = 'ingreso'"
    ).fetchone()['total']

    egresos = conn.execute(
        "SELECT COALESCE(SUM(monto), 0) as total FROM transacciones WHERE tipo = 'egreso'"
    ).fetchone()['total']

    ordenes_pendientes = conn.execute(
        "SELECT COUNT(*) as count FROM ordenes_compra WHERE estado = 'pendiente'"
    ).fetchone()['count']

    facturas_pendientes = conn.execute(
        "SELECT COALESCE(SUM(total), 0) as total FROM facturas WHERE estado = 'pendiente'"
    ).fetchone()['total']

    valor_inventario = conn.execute(
        'SELECT COALESCE(SUM(precio_costo * stock), 0) as total FROM productos'
    ).fetchone()['total']

    conn.close()

    return success_response({
        'total_ventas': total_ventas,
        'num_facturas': num_facturas,
        'total_productos': total_productos,
        'stock_bajo': stock_bajo,
        'total_clientes': total_clientes,
        'total_proveedores': total_proveedores,
        'total_empleados': total_empleados,
        'ingresos': ingresos,
        'egresos': egresos,
        'balance_financiero': ingresos - egresos,
        'ordenes_pendientes': ordenes_pendientes,
        'facturas_pendientes': facturas_pendientes,
        'valor_inventario': valor_inventario
    })


@reportes_bp.route('/dashboard/graficos', methods=['GET'])
@token_required
def dashboard_graficos():
    """Retorna datos para gráficos del dashboard."""
    conn = get_db()

    # Ventas por mes (últimos 6 meses)
    ventas_mes = conn.execute('''
        SELECT strftime('%Y-%m', fecha) as mes, SUM(total) as total, COUNT(*) as cantidad
        FROM facturas WHERE estado != 'anulada'
        GROUP BY mes ORDER BY mes DESC LIMIT 6
    ''').fetchall()

    # Top 5 productos
    top_productos = conn.execute('''
        SELECT p.nombre, SUM(fi.cantidad) as vendidos, SUM(fi.subtotal) as ingresos
        FROM items_factura fi
        JOIN productos p ON fi.producto_id = p.id
        JOIN facturas f ON fi.factura_id = f.id
        WHERE f.estado != 'anulada'
        GROUP BY p.id ORDER BY vendidos DESC LIMIT 5
    ''').fetchall()

    # Productos por categoría
    por_categoria = conn.execute('''
        SELECT categoria, COUNT(*) as cantidad, SUM(stock) as stock_total
        FROM productos GROUP BY categoria
    ''').fetchall()

    # Ingresos vs egresos por mes
    finanzas_mes = conn.execute('''
        SELECT strftime('%Y-%m', fecha) as mes, tipo, SUM(monto) as total
        FROM transacciones
        GROUP BY mes, tipo ORDER BY mes DESC LIMIT 12
    ''').fetchall()

    conn.close()

    return success_response({
        'ventas_por_mes': rows_to_list(ventas_mes),
        'top_productos': rows_to_list(top_productos),
        'productos_por_categoria': rows_to_list(por_categoria),
        'finanzas_por_mes': rows_to_list(finanzas_mes)
    })


@reportes_bp.route('/reportes/pdf', methods=['GET'])
@token_required
def exportar_pdf():
    """Genera un reporte PDF del sistema."""
    try:
        from reportlab.lib.pagesizes import letter
        from reportlab.pdfgen import canvas
        from reportlab.lib.units import inch
    except ImportError:
        return success_response({'error': 'reportlab no instalado'}, status_code=500)

    tipo = request.args.get('tipo', 'resumen')
    conn = get_db()

    buffer = io.BytesIO()
    c = canvas.Canvas(buffer, pagesize=letter)
    width, height = letter

    # Encabezado
    c.setFont('Helvetica-Bold', 18)
    c.drawString(1 * inch, height - 1 * inch, 'EasyERP - Reporte del Sistema')
    c.setFont('Helvetica', 10)
    c.drawString(1 * inch, height - 1.3 * inch, f'Generado: {datetime.now().strftime("%Y-%m-%d %H:%M")}')
    c.drawString(1 * inch, height - 1.5 * inch, f'Tipo: {tipo}')

    y = height - 2 * inch

    if tipo == 'inventario':
        c.setFont('Helvetica-Bold', 14)
        c.drawString(1 * inch, y, 'Inventario de Productos')
        y -= 0.3 * inch
        c.setFont('Helvetica', 9)
        productos = conn.execute('SELECT * FROM productos ORDER BY nombre').fetchall()
        for p in productos:
            if y < 1 * inch:
                c.showPage()
                y = height - 1 * inch
            c.drawString(1 * inch, y, f'{p["codigo"]} | {p["nombre"]} | Stock: {p["stock"]} | ${p["precio_venta"]:.2f}')
            y -= 0.2 * inch

    elif tipo == 'ventas':
        c.setFont('Helvetica-Bold', 14)
        c.drawString(1 * inch, y, 'Reporte de Ventas')
        y -= 0.3 * inch
        c.setFont('Helvetica', 9)
        facturas = conn.execute('''
            SELECT f.*, c.nombre as cliente FROM facturas f
            JOIN clientes c ON f.cliente_id = c.id
            ORDER BY f.fecha DESC LIMIT 50
        ''').fetchall()
        for f in facturas:
            if y < 1 * inch:
                c.showPage()
                y = height - 1 * inch
            c.drawString(1 * inch, y, f'#{f["id"]} | {f["fecha"]} | {f["cliente"]} | ${f["total"]:.2f} | {f["estado"]}')
            y -= 0.2 * inch

    else:
        # Resumen general
        c.setFont('Helvetica-Bold', 14)
        c.drawString(1 * inch, y, 'Resumen General')
        y -= 0.4 * inch
        c.setFont('Helvetica', 10)

        stats = [
            ('Total Productos', conn.execute('SELECT COUNT(*) as c FROM productos').fetchone()['c']),
            ('Total Clientes', conn.execute('SELECT COUNT(*) as c FROM clientes').fetchone()['c']),
            ('Total Empleados', conn.execute('SELECT COUNT(*) as c FROM empleados').fetchone()['c']),
            ('Total Ventas', f'${conn.execute("SELECT COALESCE(SUM(total),0) as t FROM facturas WHERE estado != \'anulada\'").fetchone()["t"]:.2f}'),
        ]
        for label, value in stats:
            c.drawString(1 * inch, y, f'{label}: {value}')
            y -= 0.25 * inch

    conn.close()
    c.save()
    buffer.seek(0)

    return send_file(
        buffer,
        as_attachment=True,
        download_name=f'reporte_{tipo}_{datetime.now().strftime("%Y%m%d")}.pdf',
        mimetype='application/pdf'
    )


@reportes_bp.route('/reportes/excel', methods=['GET'])
@token_required
def exportar_excel():
    """Exporta datos a Excel."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
    except ImportError:
        return success_response({'error': 'openpyxl no instalado'}, status_code=500)

    tipo = request.args.get('tipo', 'productos')
    periodo = request.args.get('periodo', '')
    conn = get_db()

    wb = Workbook()
    ws = wb.active

    header_font = Font(bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill(start_color='2563EB', end_color='2563EB', fill_type='solid')
    header_fill_green = PatternFill(start_color='16A34A', end_color='16A34A', fill_type='solid')
    header_fill_red = PatternFill(start_color='DC2626', end_color='DC2626', fill_type='solid')
    header_fill_purple = PatternFill(start_color='7C3AED', end_color='7C3AED', fill_type='solid')
    header_fill_orange = PatternFill(start_color='D97706', end_color='D97706', fill_type='solid')
    title_font = Font(bold=True, size=14, color='1E293B')
    subtitle_font = Font(bold=True, size=11, color='475569')
    money_format = '#,##0.00'
    thin_border = Border(
        left=Side(style='thin', color='E2E8F0'),
        right=Side(style='thin', color='E2E8F0'),
        top=Side(style='thin', color='E2E8F0'),
        bottom=Side(style='thin', color='E2E8F0'),
    )

    def style_header(ws, row_num, fill=header_fill):
        for cell in ws[row_num]:
            cell.font = header_font
            cell.fill = fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border

    def style_data(ws, start_row, end_row, money_cols=None):
        for row in ws.iter_rows(min_row=start_row, max_row=end_row):
            for cell in row:
                cell.border = thin_border
                cell.alignment = Alignment(vertical='center')
                if money_cols and cell.column in money_cols:
                    cell.number_format = money_format

    def auto_width(ws):
        for col in ws.columns:
            max_len = 0
            col_letter = None
            for cell in col:
                if hasattr(cell, 'column_letter'):
                    col_letter = cell.column_letter
                try:
                    if cell.value:
                        max_len = max(max_len, len(str(cell.value)))
                except:
                    pass
            if col_letter:
                ws.column_dimensions[col_letter].width = min(max_len + 4, 35)

    # ═══════════════════════════════════════════════
    #  REPORTE: NÓMINA DETALLADA (4 hojas)
    # ═══════════════════════════════════════════════
    if tipo == 'nomina_detallada':
        ws.title = 'Nómina'

        # ── HOJA 1: NÓMINA ──
        ws.append(['REPORTE DE NÓMINA DETALLADA'])
        ws.merge_cells('A1:P1')
        ws['A1'].font = title_font
        ws.append([f'Periodo: {periodo or "Todos"}', '', '', f'Generado: {datetime.now().strftime("%Y-%m-%d %H:%M")}'])
        ws['A2'].font = subtitle_font
        ws.append([])

        headers_nomina = [
            'Empleado', 'Cédula', 'Departamento', 'Puesto', 'Periodo',
            'Salario Bruto', 'Horas Extras ($)', 'Bonificaciones', 'Total Ingresos',
            'IESS (9.45%)', 'ISR', 'Desc. Faltas', 'Otras Ded.',
            'Total Deducciones', 'SALARIO NETO', 'Estado', 'Fecha Pago'
        ]
        ws.append(headers_nomina)
        style_header(ws, 4)

        query = '''
            SELECT n.*, e.nombre as empleado_nombre, e.cedula, e.puesto, e.departamento
            FROM nomina n JOIN empleados e ON n.empleado_id = e.id
        '''
        params = []
        if periodo:
            query += ' WHERE n.periodo = ?'
            params.append(periodo)
        query += ' ORDER BY n.periodo DESC, e.nombre'

        nominas = conn.execute(query, params).fetchall()
        nominas = conn.execute(query, params).fetchall()
        row_start = 5
        for n in nominas:
            ws.append([
                n['empleado_nombre'], n['cedula'], n['departamento'], n['puesto'], n['periodo'],
                n['salario_bruto'], n['horas_extras_monto'], n['bonificaciones'], n['total_ingresos'],
                n['deduccion_iess'], n['deduccion_isr'],
                n['desc_faltas_injustificadas'], n['otras_deducciones'],
                n['total_deducciones'], n['salario_neto'], n['estado'],
                n['fecha_pago'] or ''
            ])

        row_end = row_start + len(nominas) - 1
        if nominas:
            style_data(ws, row_start, row_end, money_cols=[6, 7, 8, 9, 10, 11, 12, 13, 14, 15])

            # Fila de totales
            ws.append([])
            total_row = row_end + 2
            ws.append([
                'TOTALES', '', '', '', '',
                f'=SUM(F{row_start}:F{row_end})',
                f'=SUM(G{row_start}:G{row_end})',
                f'=SUM(H{row_start}:H{row_end})',
                f'=SUM(I{row_start}:I{row_end})',
                f'=SUM(J{row_start}:J{row_end})',
                f'=SUM(K{row_start}:K{row_end})',
                f'=SUM(L{row_start}:L{row_end})',
                f'=SUM(M{row_start}:M{row_end})',
                f'=SUM(N{row_start}:N{row_end})',
                f'=SUM(O{row_start}:O{row_end})',
            ])
            for cell in ws[total_row]:
                cell.font = Font(bold=True, size=11)
                cell.number_format = money_format

        auto_width(ws)

        # ── HOJA 2: ASISTENCIA ──
        ws2 = wb.create_sheet('Asistencia')
        ws2.append(['REPORTE DE ASISTENCIA'])
        ws2.merge_cells('A1:J1')
        ws2['A1'].font = title_font
        ws2.append([f'Periodo: {periodo or "Todos"}'])
        ws2['A2'].font = subtitle_font
        ws2.append([])

        headers_asist = [
            'Empleado', 'Fecha', 'Entrada', 'Salida', 'Horas Trabajadas',
            'Horas Extra', 'Tipo', 'Observación'
        ]
        ws2.append(headers_asist)
        style_header(ws2, 4, header_fill_green)

        query_asist = '''
            SELECT a.*, e.nombre as empleado_nombre
            FROM asistencia a JOIN empleados e ON a.empleado_id = e.id
        '''
        params_asist = []
        if periodo:
            query_asist += " WHERE strftime('%Y-%m', a.fecha) = ?"
            params_asist.append(periodo)
        query_asist += ' ORDER BY a.fecha DESC, e.nombre'

        asistencias = conn.execute(query_asist, params_asist).fetchall()
        tipo_labels = {
            'normal': 'Normal', 'tardanza': 'Tardanza',
            'falta_justificada': 'Falta Justificada', 'falta_injustificada': 'Falta Injustificada'
        }
        r_start = 5
        for a in asistencias:
            ws2.append([
                a['empleado_nombre'], a['fecha'], a['entrada'] or '-', a['salida'] or '-',
                a['horas_trabajadas'], a['horas_extra'],
                tipo_labels.get(a['tipo'], a['tipo']),
                a['observacion'] or ''
            ])

        if asistencias:
            style_data(ws2, r_start, r_start + len(asistencias) - 1)

        # Resumen por empleado
        ws2.append([])
        ws2.append(['RESUMEN POR EMPLEADO'])
        sum_row = r_start + len(asistencias) + 1
        ws2.cell(row=sum_row, column=1).font = subtitle_font
        ws2.append([])

        resumen_headers = ['Empleado', 'Días Trabajados', 'Tardanzas', 'Faltas Just.', 'Faltas Injust.', 'Total Horas', 'Total Hrs Extra']
        ws2.append(resumen_headers)
        style_header(ws2, sum_row + 2, header_fill_orange)

        resumen_asist = conn.execute('''
            SELECT e.nombre,
                SUM(CASE WHEN a.tipo IN ('normal', 'tardanza') THEN 1 ELSE 0 END) as dias_trabajados,
                SUM(CASE WHEN a.tipo = 'tardanza' THEN 1 ELSE 0 END) as tardanzas,
                SUM(CASE WHEN a.tipo = 'falta_justificada' THEN 1 ELSE 0 END) as faltas_just,
                SUM(CASE WHEN a.tipo = 'falta_injustificada' THEN 1 ELSE 0 END) as faltas_injust,
                COALESCE(SUM(a.horas_trabajadas), 0) as total_horas,
                COALESCE(SUM(a.horas_extra), 0) as total_extras
            FROM asistencia a JOIN empleados e ON a.empleado_id = e.id
        ''' + (" WHERE strftime('%Y-%m', a.fecha) = ?" if periodo else '') +
            ' GROUP BY e.id ORDER BY e.nombre',
            params_asist
        ).fetchall()

        res_start = sum_row + 3
        for r in resumen_asist:
            ws2.append([r['nombre'], r['dias_trabajados'], r['tardanzas'],
                        r['faltas_just'], r['faltas_injust'], r['total_horas'], r['total_extras']])
        if resumen_asist:
            style_data(ws2, res_start, res_start + len(resumen_asist) - 1)

        auto_width(ws2)

        # ── HOJA 3: INCIDENCIAS ──
        ws3 = wb.create_sheet('Incidencias')
        ws3.append(['REPORTE DE INCIDENCIAS'])
        ws3.merge_cells('A1:H1')
        ws3['A1'].font = title_font
        ws3.append([f'Periodo: {periodo or "Todos"}'])
        ws3['A2'].font = subtitle_font
        ws3.append([])

        headers_inc = ['Empleado', 'Puesto', 'Fecha', 'Tipo', 'Descripción', 'Días Ausencia', 'Justificada', 'Documento']
        ws3.append(headers_inc)
        style_header(ws3, 4, header_fill_red)

        query_inc = '''
            SELECT i.*, e.nombre as empleado_nombre, e.puesto
            FROM incidencias i JOIN empleados e ON i.empleado_id = e.id
        '''
        params_inc = []
        if periodo:
            query_inc += " WHERE strftime('%Y-%m', i.fecha) = ?"
            params_inc.append(periodo)
        query_inc += ' ORDER BY i.fecha DESC'

        tipo_inc_labels = {
            'medica': 'Médica', 'personal': 'Personal', 'disciplinaria': 'Disciplinaria',
            'accidente': 'Accidente', 'otra': 'Otra'
        }

        incidencias = conn.execute(query_inc, params_inc).fetchall()
        i_start = 5
        for i in incidencias:
            ws3.append([
                i['empleado_nombre'], i['puesto'], i['fecha'],
                tipo_inc_labels.get(i['tipo'], i['tipo']),
                i['descripcion'], i['dias_ausencia'],
                'Sí' if i['justificada'] else 'No',
                i['documento_soporte'] or '-'
            ])

        if incidencias:
            style_data(ws3, i_start, i_start + len(incidencias) - 1)
        auto_width(ws3)

        # ── HOJA 4: HORAS EXTRAS ──
        ws4 = wb.create_sheet('Horas Extras')
        ws4.append(['REPORTE DE HORAS EXTRAS'])
        ws4.merge_cells('A1:H1')
        ws4['A1'].font = title_font
        ws4.append([f'Periodo: {periodo or "Todos"}'])
        ws4['A2'].font = subtitle_font
        ws4.append([])

        headers_he = ['Empleado', 'Puesto', 'Fecha', 'Horas', 'Tipo', 'Monto ($)', 'Aprobado', 'Observación']
        ws4.append(headers_he)
        style_header(ws4, 4, header_fill_purple)

        query_he = '''
            SELECT he.*, e.nombre as empleado_nombre, e.puesto
            FROM horas_extras he JOIN empleados e ON he.empleado_id = e.id
        '''
        params_he = []
        if periodo:
            query_he += " WHERE strftime('%Y-%m', he.fecha) = ?"
            params_he.append(periodo)
        query_he += ' ORDER BY he.fecha DESC'

        tipo_he_labels = {'normal': '1.5x Normal', 'doble': '2x Doble', 'triple': '3x Triple'}
        horas_extras = conn.execute(query_he, params_he).fetchall()
        he_start = 5
        for he in horas_extras:
            ws4.append([
                he['empleado_nombre'], he['puesto'], he['fecha'], he['horas'],
                tipo_he_labels.get(he['tipo'], he['tipo']),
                he['monto'],
                'Sí' if he['aprobado'] else 'No',
                he['observacion'] or ''
            ])

        if horas_extras:
            style_data(ws4, he_start, he_start + len(horas_extras) - 1, money_cols=[6])

            # Total
            ws4.append([])
            ws4.append(['TOTAL', '', '', f'=SUM(D{he_start}:D{he_start + len(horas_extras) - 1})',
                         '', f'=SUM(F{he_start}:F{he_start + len(horas_extras) - 1})'])
            total_he_row = he_start + len(horas_extras) + 1
            for cell in ws4[total_he_row]:
                cell.font = Font(bold=True)
                if cell.column in [4, 6]:
                    cell.number_format = money_format

        auto_width(ws4)

    # ═══════════════════════════════════════════════
    #  REPORTES EXISTENTES (sin cambios)
    # ═══════════════════════════════════════════════
    elif tipo == 'productos':
        ws.title = 'Productos'
        headers = ['ID', 'Código', 'Nombre', 'Categoría', 'Stock', 'Precio Costo', 'Precio Venta']
        ws.append(headers)
        style_header(ws, 1)
        for p in conn.execute('SELECT * FROM productos ORDER BY nombre').fetchall():
            ws.append([p['id'], p['codigo'], p['nombre'], p['categoria'],
                       p['stock'], p['precio_costo'], p['precio_venta']])
        if ws.max_row > 1:
            style_data(ws, 2, ws.max_row, money_cols=[6, 7])
        auto_width(ws)

    elif tipo == 'ventas':
        ws.title = 'Ventas'
        headers = ['ID', 'Fecha', 'Cliente', 'Total', 'Estado']
        ws.append(headers)
        style_header(ws, 1)
        for f in conn.execute('''
            SELECT f.id, f.fecha, c.nombre, f.total, f.estado
            FROM facturas f JOIN clientes c ON f.cliente_id = c.id
            ORDER BY f.fecha DESC
        ''').fetchall():
            ws.append([f['id'], f['fecha'], f['nombre'], f['total'], f['estado']])
        if ws.max_row > 1:
            style_data(ws, 2, ws.max_row, money_cols=[4])
        auto_width(ws)

    elif tipo == 'empleados':
        ws.title = 'Empleados'
        headers = ['ID', 'Nombre', 'Cédula', 'Departamento', 'Puesto', 'Tipo Contrato',
                   'Horas/Semana', 'Salario', 'Fecha Ingreso', 'Estado']
        ws.append(headers)
        style_header(ws, 1)
        for e in conn.execute('SELECT * FROM empleados ORDER BY nombre').fetchall():
            ws.append([e['id'], e['nombre'], e['cedula'], e['departamento'], e['puesto'],
                       e['tipo_contrato'], e['horas_semanales'], e['salario'],
                       e['fecha_ingreso'], e['estado']])
        if ws.max_row > 1:
            style_data(ws, 2, ws.max_row, money_cols=[8])
        auto_width(ws)

    else:
        ws.title = 'Transacciones'
        headers = ['ID', 'Tipo', 'Descripción', 'Monto', 'Fecha']
        ws.append(headers)
        style_header(ws, 1)
        for t in conn.execute('SELECT * FROM transacciones ORDER BY fecha DESC').fetchall():
            ws.append([t['id'], t['tipo'], t['descripcion'], t['monto'], t['fecha']])
        if ws.max_row > 1:
            style_data(ws, 2, ws.max_row, money_cols=[4])
        auto_width(ws)

    conn.close()

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return send_file(
        buffer,
        as_attachment=True,
        download_name=f'export_{tipo}_{datetime.now().strftime("%Y%m%d")}.xlsx',
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
